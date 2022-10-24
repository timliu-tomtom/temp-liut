#!/usr/bin/env python3

# Copyright (C) 2021 TomTom NV. All rights reserved.
#
# This software is the proprietary copyright of TomTom NV and its subsidiaries and may be
# used for internal evaluation purposes or commercial use strictly subject to separate
# license agreement between you and TomTom NV. If you are the licensee, you are only permitted
# to use this software in accordance with the terms of your license agreement. If you are
# not the licensee, you are not authorized to use this software in any manner and should
# immediately return or destroy it.

""" PO - DITA Converter

Usage:
  po-dita-converter po2dita <template> <output>
  po-dita-converter dita2po <input> <output>
  po-dita-converter validate <input>
  po-dita-converter (-h | --help)
  po-dita-converter --version
  po-dita-converter

Commands:
  po2dita              to convert from po to dita file
  dita2po              to convert from dita to po file
  validate             check if file has valid xml schema

Options:
  -h --help            Show this screen.
  --version            Show version.
  template             file used for en-GB template strings
  input                path to xml file
  output               path to po file
"""

# build in
import sys
import os
import openpyxl as xl

# extern
from docopt import docopt
from lxml import etree

# internal
import src.convert_po_to_dita as po_to_dita
import src.convert_dita_to_po as dita_to_po


def main():
    wb = xl.load_workbook('test.xlsx')
    print(wb.sheetnames)
    sheet = wb['Blad1']
    c = sheet['A1']
    print("A1 " + c.value)
    c = sheet['A2']
    print("A2 " + c.value)

    #sheet['A2'] = "高速道路を {distance} 走行します。"
    sheet.cell(row=2, column=1, value="高速道路を {distance} 走行します。")

    en_tree = etree.parse('en-GB.xml')
    en_root = en_tree.getroot()

    ja_tree = etree.parse('ja-JP.xml')
    ja_root = ja_tree.getroot()

    updated_ja_tree = etree.parse('updated-ja-JP.xml')
    updated_ja_root = updated_ja_tree.getroot()

    row = 2
    for en_child in en_tree.iter("uicontrol"):
        id = en_child.attrib.get('id')
        if not id:
            continue

        en_message = en_child.text
        if not en_message:
            continue

        ja_message = ""
        for ja_child in ja_tree.iter("uicontrol"):
            if ja_child.attrib.get('id') == id:
                ja_message = ja_child.text
                break

        updated_ja_message = ""
        for updated_ja_child in updated_ja_tree.iter("uicontrol"):
            if updated_ja_child.attrib.get('id') == id:
                updated_ja_message = updated_ja_child.text
                break

        if ja_message != updated_ja_message:
            sheet.cell(row=row, column=1, value=en_message)
            sheet.cell(row=row, column=2, value=ja_message)
            sheet.cell(row=row, column=3, value=updated_ja_message)
            sheet.cell(row=row, column=4, value=id)
            row = row + 1

    wb.save('output.xlsx')

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # execute only if run as a script
    main()
